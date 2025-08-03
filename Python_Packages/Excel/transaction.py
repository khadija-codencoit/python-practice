import openpyxl

#wb = openpyxl.workbook()
wb = openpyxl.load_workbook("xcel-files/transaction (1).xlsx")
print(wb.sheetnames)

sheet = wb['Sheet1']
cell = sheet['a1']

print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

print(sheet.max_column)
print(sheet.max_row)

column = sheet["a"]
print(column)
cells = sheet['a:c']
print(cells)
sheet.append([111,222,333])
# wb.save("xcel-files/transaction2.xlsx")

box = sheet['a1']
for row in range(1,sheet.max_row+1):
    for colum in range(1,sheet.max_column+1):
        cell = sheet.cell(row,colum)
        print(cell.value)